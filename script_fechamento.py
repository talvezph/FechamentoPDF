import os
import pandas as pd
import pdfplumber
from collections import defaultdict
from datetime import datetime
import difflib
import re
import unicodedata
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
import configparser
import logging
import argparse

# Lista para armazenar mensagens de erro e aviso
error_report_messages = []

# Configuração de logging
class CustomFormatter(logging.Formatter):
    def format(self, record):
        msg = super().format(record)
        if record.levelno in [logging.WARNING, logging.ERROR]:
            error_report_messages.append(msg)
        return msg

logger = logging.getLogger()
logger.setLevel(logging.INFO)

console_handler = logging.StreamHandler()
console_handler.setFormatter(CustomFormatter("%(asctime)s - %(levelname)s - %(message)s"))
logger.addHandler(console_handler)

# Carrega configurações do arquivo config.ini
config = configparser.ConfigParser()
config.read("config.ini")

# Argumentos de linha de comando
parser = argparse.ArgumentParser(description="Processa PDFs de motoristas para gerar um fechamento em Excel.")
parser.add_argument("--pdfs_folder", type=str, help="Caminho para a pasta contendo os PDFs.")
parser.add_argument("--type_sheet", type=str, help="Caminho para a planilha de tipos de veículos.")
parser.add_argument("--output_excel", type=str, help="Nome do arquivo Excel de saída.")
parser.add_argument("--error_report", type=str, default="error_report.log", help="Nome do arquivo para o relatório de erros.")
args = parser.parse_args()

# Caminhos (prioriza argumentos de linha de comando)
PASTA_PDFS = args.pdfs_folder if args.pdfs_folder else config["Paths"]["pdfs_folder"]
PLANILHA_TIPO = args.type_sheet if args.type_sheet else config["Paths"]["type_sheet"]
SAIDA_EXCEL = args.output_excel if args.output_excel else config["Paths"]["output_excel"]
ERROR_REPORT_FILE = args.error_report

# Valores fixos
VALOR_ENTREGA = float(config["Values"]["delivery_value"])
BONUS_DIARIO = float(config["Values"]["daily_bonus"])

# Expressões regulares
regex_data = re.compile(r"\b\d{2}/\d{2}/\d{4}\b")
regex_valor = re.compile(r"R\$\s*([\d.,]+)")

# Normalização de texto
def normalize(texto):
    nfkd = unicodedata.normalize("NFKD", texto)
    return u"".join([c for c in nfkd if not unicodedata.combining(c)]).lower()

# Leitura da planilha de tipos de veículos
try:
    df_veiculos = pd.read_excel(PLANILHA_TIPO)
    df_veiculos.columns = df_veiculos.columns.str.strip().str.lower()

    # Validação: Verificar se as colunas essenciais existem
    required_cols_name = ["nome do motorista", "motorista", "nome"]
    required_cols_diaria = ["diária combinada", "diaria combinada", "diaria"]
    
    found_name_col = any(col in df_veiculos.columns for col in required_cols_name)
    found_diaria_col = any(col in df_veiculos.columns for col in required_cols_diaria)

    if not found_name_col:
        logging.error(f"Erro: Nenhuma das colunas de nome de motorista ({required_cols_name}) encontrada na planilha {PLANILHA_TIPO}.")
        exit(1)
    if not found_diaria_col:
        logging.error(f"Erro: Nenhuma das colunas de diária combinada ({required_cols_diaria}) encontrada na planilha {PLANILHA_TIPO}.")
        exit(1)

    tipo_colunas = [col for col in df_veiculos.columns if "tipo" in col]
    coluna_tipo = tipo_colunas[0] if tipo_colunas else None
    if not coluna_tipo:
        logging.warning(f"Aviso: Nenhuma coluna contendo \"tipo\" foi encontrada na planilha {PLANILHA_TIPO}. O tipo de veículo não será registrado.")

    diarios_info = {}
    for idx, linha in df_veiculos.iterrows():
        nome = None
        for col_n in required_cols_name:
            if col_n in df_veiculos.columns:
                nome = linha.get(col_n)
                break
        
        if pd.isna(nome) or str(nome).strip() == "":
            logging.warning(f"Aviso: Nome de motorista inválido ou vazio na linha {idx + 2} da planilha {PLANILHA_TIPO}. Pulando esta linha.")
            continue
        nome_str = str(nome).strip().upper()
        
        diaria_valor = 0
        for col_d in required_cols_diaria:
            if col_d in df_veiculos.columns:
                diaria_valor = linha.get(col_d, 0) or 0
                break
        
        if not isinstance(diaria_valor, (int, float)) or diaria_valor < 0:
            logging.warning(f"Aviso: Valor de diária inválido ({diaria_valor}) para o motorista {nome_str} na linha {idx + 2} da planilha {PLANILHA_TIPO}. Usando 0.")
            diaria_valor = 0

        tipo_valor = linha.get(coluna_tipo) if coluna_tipo else None
        if pd.isna(tipo_valor) or str(tipo_valor).strip() == "":
            logging.warning(f"Aviso: Tipo de veículo inválido ou vazio para o motorista {nome_str} na linha {idx + 2} da planilha {PLANILHA_TIPO}. Usando \"N/A\".")
            tipo_valor = "N/A"

        diarios_info[nome_str] = {"diaria": diaria_valor, "tipo": tipo_valor}

    if not diarios_info:
        logging.error("Nenhum motorista válido encontrado na planilha de veículos. Verifique os nomes e colunas.")
        exit(1)
except FileNotFoundError:
    logging.error(f"Erro: A planilha de tipos de veículos {PLANILHA_TIPO} não foi encontrada.")
    exit(1)
except Exception as e:
    logging.error(f"Erro ao ler a planilha de tipos de veículos: {e}")
    exit(1)

# Busca nome aproximado
def encontrar_nome_aproximado(nome_pdf):
    # Normaliza o nome do PDF para comparação
    nome_pdf_normalizado = normalize(nome_pdf)
    
    # Normaliza os nomes dos motoristas na planilha para comparação
    nomes_planilha_normalizados = {normalize(nome): nome for nome in diarios_info.keys()}
    
    # Ajuste o cutoff para um valor mais flexível, por exemplo, 0.6 ou 0.5
    matches = difflib.get_close_matches(nome_pdf_normalizado, list(nomes_planilha_normalizados.keys()), n=1, cutoff=0.5) 
    
    if matches:
        return nomes_planilha_normalizados[matches[0]]
    else:
        return None

# Extrai dados do PDF: nome, entregas, acréscimos e bônus pagos via análise linha a linha
def extrair_dados_pdf(caminho_pdf):
    entregas_por_dia = defaultdict(lambda: {"entregues": 0, "insucessos": 0})
    acres_por_data = defaultdict(float)
    bonus_pago_dates = set()
    nome_motorista = None

    logging.info(f"Processando PDF: {os.path.basename(caminho_pdf)}")
    try:
        with pdfplumber.open(caminho_pdf) as pdf:
            full_text = ""
            all_tables = []
            for pagina in pdf.pages:
                full_text += pagina.extract_text() or ""
                all_tables.extend(pagina.extract_tables())

            # Processa todas as tabelas uma única vez
            for table in all_tables:
                for row in table:
                    data_match = regex_data.search(str(row))
                    valor_match = regex_valor.search(str(row))
                    if data_match and valor_match:
                        try:
                            data = datetime.strptime(data_match.group(), "%d/%m/%Y").date()
                            valor_str = valor_match.group(1).replace(".", "").replace(",", ".")
                            acres_por_data[data] += float(valor_str)
                            logging.info(f"  Acréscimo de tabela encontrado: Data {data}, Valor {valor_str}")
                        except ValueError:
                            logging.warning(f"  Valor ou data inválida em tabela de acréscimo: {row} no PDF {os.path.basename(caminho_pdf)}")
                        except Exception as e:
                            logging.warning(f"  Erro ao processar linha de tabela de acréscimo: {e} na linha: {row} no PDF {os.path.basename(caminho_pdf)}")

            linhas_texto = full_text.split("\n")
            remuneracoes_section = False

            for i, linha in enumerate(linhas_texto):
                low = normalize(linha)
                # Identifica nome do motorista
                if not nome_motorista and "motorista:" in low:
                    parts = linha.split(":")
                    if len(parts) > 1:
                        nome_motorista = parts[-1].strip().upper()
                        logging.info(f"  Motorista identificado: {nome_motorista}")
                        continue
                # Detecta início de "Remunerações Diárias"
                if "remuneracoes diarias" in low or "remunerações diárias" in low:
                    remuneracoes_section = True
                    continue

                # Se estiver na seção de remunerações, busca data + 30,00 e pula demais
                if remuneracoes_section:
                    if regex_data.search(linha) and "30,00" in linha:
                        m_data = regex_data.search(linha)
                        try:
                            data = datetime.strptime(m_data.group(), "%d/%m/%Y").date()
                            bonus_pago_dates.add(data)
                        except ValueError:
                            logging.warning(f"  Data inválida encontrada no bônus: {linha} no PDF {os.path.basename(caminho_pdf)}")
                        except Exception as e:
                            logging.warning(f"  Erro ao processar bônus: {e} na linha: {linha} no PDF {os.path.basename(caminho_pdf)}")
                    if low.strip() == "" or "coletas/entregas" in low:
                        remuneracoes_section = False
                    continue
                
                # Fora das seções de remunerações e acréscimos, identifica entregas e acréscimos de linha
                if regex_data.search(linha):
                    m_data = regex_data.search(linha)
                    try:
                        data = datetime.strptime(m_data.group(), "%d/%m/%Y").date()
                    except ValueError:
                        logging.warning(f"  Data inválida encontrada: {linha} no PDF {os.path.basename(caminho_pdf)}")
                        continue
                    except Exception as e:
                        logging.warning(f"  Erro ao processar data: {e} na linha: {linha} no PDF {os.path.basename(caminho_pdf)}")
                        continue
                    # Entregas/Insucessos: procura sim/nao
                    if "sim" in low or "nao" in low:
                        status = "Sim" if "sim" in low else "Não"
                        if status == "Sim":
                            entregas_por_dia[data]["entregues"] += 1
                        else:
                            entregas_por_dia[data]["insucessos"] += 1
                    # Acréscimos: tem R$ mas não sim/nao (para acréscimos que não estão em tabela)
                    elif regex_valor.search(linha) and not ("sim" in low or "nao" in low):
                        m_val = regex_valor.search(linha)
                        if m_val:
                            # Remove pontos de milhar e substitui vírgula por ponto decimal
                            valor_str = m_val.group(1).replace(".", "").replace(",", ".")
                            try:
                                acres_por_data[data] += float(valor_str)
                            except ValueError:
                                logging.warning(f"  Valor inválido encontrado no acréscimo: {linha} no PDF {os.path.basename(caminho_pdf)}")
                            except Exception as e:
                                logging.warning(f"  Erro ao processar acréscimo: {e} na linha: {linha} no PDF {os.path.basename(caminho_pdf)}")
    except pdfplumber.PDFSyntaxError:
        logging.error(f"Erro de sintaxe no PDF: {os.path.basename(caminho_pdf)}. O arquivo pode estar corrompido ou não é um PDF válido.")
        return None, defaultdict(lambda: {"entregues": 0, "insucessos": 0}), defaultdict(float), set()
    except Exception as e:
        logging.error(f"Erro inesperado ao extrair dados do PDF {os.path.basename(caminho_pdf)}: {e}")
        return None, defaultdict(lambda: {"entregues": 0, "insucessos": 0}), defaultdict(float), set()
    return nome_motorista, entregas_por_dia, acres_por_data, bonus_pago_dates

# Calcula fechamento do motorista
def calcular_fechamento(nome_motorista, entregas_por_dia, acres_por_data, bonus_pago_dates):
    nome_upper = nome_motorista.strip().upper()
    nome_final = encontrar_nome_aproximado(nome_upper)
    if not nome_final:
        logging.warning(f"  Nome não encontrado na planilha de veículos para: {nome_motorista}")
        return None, pd.DataFrame()
    info = diarios_info[nome_final]
    diaria, tipo_veiculo = info["diaria"], info["tipo"]

    registros = []
    total_rotas = 0.0
    todas_datas = set(entregas_por_dia.keys()) | set(acres_por_data.keys())

    for data in sorted(todas_datas):
        entregues = entregas_por_dia.get(data, {"entregues": 0})["entregues"]
        insucessos = entregas_por_dia.get(data, {"insucessos": 0})["insucessos"]
        valor_entregas = entregues * VALOR_ENTREGA
        descontos = insucessos * VALOR_ENTREGA
        acres_pago = acres_por_data.get(data, 0.0)
        acrescimo_calculado = max(0.0, diaria - valor_entregas - descontos)
        total_dia = valor_entregas + acres_pago - descontos
        total_rotas += total_dia
        bonus_val = BONUS_DIARIO if data in bonus_pago_dates else 0

        registros.append({
            "Data": data.strftime("%d/%m/%Y"),
            "Motorista": nome_final,
            "Tipo de Veículo": tipo_veiculo,
            "Entregues": entregues,
            "Insucessos": insucessos,
            "Valor Entregas": valor_entregas,
            "Descontos": descontos,
            "Acréscimo Calculado": acrescimo_calculado,
            "Acréscimo Pago": acres_pago,
            "Total Dia": total_dia,
            "Bônus": bonus_val
        })

    registros.append({
        "Data": "Total",
        "Motorista": nome_final,
        "Tipo de Veículo": tipo_veiculo,
        "Entregues": sum(v["entregues"] for v in entregas_por_dia.values()),
        "Insucessos": sum(v["insucessos"] for v in entregas_por_dia.values()),
        "Valor Entregas": sum(v["entregues"] for v in entregas_por_dia.values()) * VALOR_ENTREGA,
        "Descontos": sum(v["insucessos"] for v in entregas_por_dia.values()) * VALOR_ENTREGA,
        "Acréscimo Calculado": 0.0,
        "Acréscimo Pago": sum(acres_por_data.values()),
        "Total Dia": total_rotas,
        "Bônus": len([d for d in todas_datas if d in bonus_pago_dates]) * BONUS_DIARIO
    })

    df = pd.DataFrame(registros)
    return nome_final, df

def main():
    # Dicionário para armazenar DataFrames por motorista
    fechamentos_consolidados = defaultdict(pd.DataFrame)

    if not os.path.exists(PASTA_PDFS):
        logging.error(f"Erro: A pasta de PDFs {PASTA_PDFS} não foi encontrada.")
        exit(1)

    # Agrupar PDFs por nome base do motorista (ignorando sufixos numéricos)
    pdf_files_grouped = defaultdict(list)
    for nome_arquivo in os.listdir(PASTA_PDFS):
        if nome_arquivo.lower().endswith(".pdf"):
            # Remove a extensão .pdf e qualquer sufixo numérico (ex: 2, 3) no final
            nome_base = re.sub(r"\d*\.pdf$", "", nome_arquivo.lower())
            nome_base = nome_base.replace(".pdf", "").strip()
            pdf_files_grouped[nome_base].append(os.path.join(PASTA_PDFS, nome_arquivo))

    for nome_base_motorista, caminhos_pdfs in pdf_files_grouped.items():
        motorista_entregas_por_dia = defaultdict(lambda: {"entregues": 0, "insucessos": 0})
        motorista_acres_por_data = defaultdict(float)
        motorista_bonus_pago_dates = set()
        motorista_nome_final = None

        for caminho_pdf in caminhos_pdfs:
            nome, entregas, acrescimos, bonus = extrair_dados_pdf(caminho_pdf)
            if nome:
                if not motorista_nome_final:
                    motorista_nome_final = nome
                
                # Consolidar entregas
                for data, info_entrega in entregas.items():
                    motorista_entregas_por_dia[data]["entregues"] += info_entrega["entregues"]
                    motorista_entregas_por_dia[data]["insucessos"] += info_entrega["insucessos"]
                
                # Consolidar acréscimos
                for data, valor_acrescimo in acrescimos.items():
                    motorista_acres_por_data[data] += valor_acrescimo
                
                # Consolidar bônus
                motorista_bonus_pago_dates.update(bonus)

        if motorista_nome_final:
            nome_final_calculado, df_fechamento = calcular_fechamento(
                motorista_nome_final, 
                motorista_entregas_por_dia, 
                motorista_acres_por_data, 
                motorista_bonus_pago_dates
            )
            if not df_fechamento.empty:
                fechamentos_consolidados[nome_final_calculado] = df_fechamento

    # Escrever no Excel
    try:
        # Cria um novo workbook
        book = Workbook()
        # Remove a aba padrão 'Sheet' se ela existir
        if 'Sheet' in book.sheetnames:
            book.remove(book['Sheet'])

        for motorista, df_novo in fechamentos_consolidados.items():
            sheet_name = motorista[:31] # Limita o nome da aba para 31 caracteres
            
            # Adiciona a nova aba
            ws = book.create_sheet(sheet_name)
            
            # Escreve o cabeçalho
            ws.append(df_novo.columns.tolist())
            
            # Escreve os dados
            for r_idx, row in df_novo.iterrows():
                ws.append(row.tolist())

            # Formatação da aba
            header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid") # LightBlue
            for cell in ws["1:1"]:
                cell.fill = header_fill
            
            # Encontrar a linha "Total" e aplicar formatação
            for row_idx in range(1, ws.max_row + 1):
                if ws.cell(row=row_idx, column=1).value == "Total":
                    for col_idx in range(1, ws.max_column + 1):
                        ws.cell(row=row_idx, column=col_idx).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid") # Yellow
                    break
        
        book.save(SAIDA_EXCEL)
        logging.info(f"Fechamento gerado com sucesso em {SAIDA_EXCEL}")
    except Exception as e:
        logging.error(f"Erro ao escrever ou formatar a planilha de saída: {e}")
        exit(1)

    # Gerar relatório de erros
    if error_report_messages:
        with open(ERROR_REPORT_FILE, "w") as f:
            for msg in error_report_messages:
                f.write(msg + "\n")
        logging.warning(f"Relatório de erros gerado em {ERROR_REPORT_FILE}")

if __name__ == "__main__":
    main()


